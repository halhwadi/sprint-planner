import json
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from .models import (
    Organization, OrganizationMember, Stream,
    Sprint, SprintMember, UserStory, Vote, StreamAssignment
)
from .permissions import (
    require_org_member, require_scrum_master, require_admin,
    require_voter, require_scrum_master_api, require_admin_api,
    require_voter_api, is_scrum_master_or_above
)

BANDWIDTH_LIMIT = 8


# ─────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────

def get_org(request):
    """Get active org for current user.
    If user belongs to multiple orgs, respects session selection."""
    active_org_id = request.session.get('active_org_id')
    if active_org_id:
        membership = OrganizationMember.objects.select_related('organization').filter(
            user=request.user, organization_id=active_org_id
        ).first()
        if membership:
            return membership.organization

    # Default — first org
    membership = OrganizationMember.objects.select_related('organization').filter(
        user=request.user
    ).first()
    return membership.organization if membership else None


def get_member(request, org):
    try:
        return SprintMember.objects.get(user=request.user, organization=org, is_active=True)
    except SprintMember.DoesNotExist:
        return None


# ─────────────────────────────────────────
# PUBLIC
# ─────────────────────────────────────────

def home(request):
    return redirect('board')

def landing(request):
    if request.user.is_authenticated:
        return redirect('sm_panel')
    return render(request, 'planner/landing.html')

def privacy_policy(request):
    return render(request, 'planner/privacy_policy.html')

def terms_of_service(request):
    return render(request, 'planner/terms_of_service.html')

def refund_policy(request):
    return render(request, 'planner/refund_policy.html')
    
# ─────────────────────────────────────────
# BOARD
# ─────────────────────────────────────────

@require_org_member
def board(request):
    org           = get_org(request)
    user_is_sm    = is_scrum_master_or_above(request.user, org)
    member        = get_member(request, org)
    sprints       = Sprint.objects.filter(organization=org)
    active_sprint = sprints.filter(is_active=True).first()
    streams       = Stream.objects.filter(organization=org)

    sprint_id       = request.GET.get('sprint')
    selected_sprint = None
    if sprint_id:
        selected_sprint = get_object_or_404(Sprint, id=sprint_id, organization=org)
    elif active_sprint:
        selected_sprint = active_sprint

    stories = UserStory.objects.filter(organization=org).prefetch_related(
        'votes', 'stream_assignments', 'stream_assignments__member'
    ).select_related('owner', 'sprint')

    if selected_sprint:
        stories = stories.filter(sprint=selected_sprint)

    all_members = SprintMember.objects.filter(
        organization=org, is_active=True
    ).select_related('user', 'stream')

    bandwidth = []
    for m in all_members:
        total = m.total_sp(sprint=selected_sprint)
        bandwidth.append({'member': m, 'total': total, 'over': total > BANDWIDTH_LIMIT})

    return render(request, 'planner/board.html', {
        'stories':         stories,
        'member':          member,
        'is_sm':           user_is_sm,
        'bandwidth':       bandwidth,
        'streams':         streams,
        'all_members':     all_members,
        'sprints':         sprints,
        'selected_sprint': selected_sprint,
        'active_sprint':   active_sprint,
        'org':             org,
    })


# ─────────────────────────────────────────
# VOTE ROOM
# ─────────────────────────────────────────

@require_org_member
def vote_room(request, us_id):
    org        = get_org(request)
    story      = get_object_or_404(UserStory, id=us_id, organization=org)
    user_is_sm = is_scrum_master_or_above(request.user, org)
    member     = get_member(request, org)
    streams    = Stream.objects.filter(organization=org)
    all_members = SprintMember.objects.filter(
        organization=org, is_active=True
    ).select_related('user', 'stream')
    fibonacci  = [1, 2, 3, 5, 8, 13]

    my_vote = None
    if member:
        try:
            my_vote = Vote.objects.get(user_story=story, member=member)
        except Vote.DoesNotExist:
            pass

    return render(request, 'planner/vote.html', {
        'story':       story,
        'member':      member,
        'is_sm':       user_is_sm,
        'fibonacci':   fibonacci,
        'my_vote':     my_vote,
        'all_members': all_members,
        'streams':     streams,
    })


def vote_status(request, us_id):
    story       = get_object_or_404(UserStory, id=us_id)
    all_members = SprintMember.objects.filter(
        organization=story.organization, is_active=True
    ).select_related('user')
    votes = {v.member_id: v.points for v in story.votes.all()}

    members_status = []
    for m in all_members:
        members_status.append({
            'id':     m.id,
            'name':   m.display_name(),
            'stream': m.stream.name if m.stream else '—',
            'voted':  m.id in votes,
            'points': votes.get(m.id) if story.voting_status == 'closed' else None,
        })

    stream_averages = []
    if story.voting_status == 'closed' and votes:
        from collections import defaultdict
        stream_votes = defaultdict(list)
        for m in all_members:
            if m.id in votes and m.stream:
                stream_votes[m.stream.name].append(votes[m.id])
        for stream, points in sorted(stream_votes.items()):
            stream_averages.append({
                'stream':  stream,
                'average': round(sum(points) / len(points), 1),
                'votes':   len(points),
            })

    return JsonResponse({
        'status':          story.voting_status,
        'members':         members_status,
        'average':         story.vote_average,
        'final_sp':        story.final_sp,
        'total_members':   all_members.count(),
        'voted_count':     len(votes),
        'stream_averages': stream_averages,
    })


@require_POST
@require_voter_api
def submit_vote(request, us_id):
    org    = get_org(request)
    story  = get_object_or_404(UserStory, id=us_id, organization=org)
    member = get_member(request, org)

    if not member:
        return JsonResponse({'error': 'You are not a sprint member'}, status=403)
    if story.voting_status != 'voting':
        return JsonResponse({'error': 'Voting not open'}, status=400)

    data   = json.loads(request.body)
    points = int(data.get('points'))
    if points not in [1, 2, 3, 5, 8, 13]:
        return JsonResponse({'error': 'Invalid points'}, status=400)

    Vote.objects.update_or_create(
        user_story=story, member=member, defaults={'points': points}
    )
    return JsonResponse({'ok': True})


# ─────────────────────────────────────────
# SM PANEL
# ─────────────────────────────────────────

@require_scrum_master
def sm_panel(request):
    org           = get_org(request)
    active_sprint = Sprint.objects.filter(organization=org, is_active=True).first()
    all_members   = SprintMember.objects.filter(
        organization=org, is_active=True
    ).select_related('user', 'stream')
    streams       = Stream.objects.filter(organization=org)

    bandwidth = []
    for m in all_members:
        total = m.total_sp(sprint=active_sprint)
        bandwidth.append({'member': m, 'total': total, 'over': total > BANDWIDTH_LIMIT})

    return render(request, 'planner/sm_panel.html', {
        'members':       all_members,
        'stories':       UserStory.objects.filter(organization=org).select_related(
                             'owner', 'sprint'
                         ).prefetch_related('stream_assignments__member'),
        'streams':       streams,
        'all_members':   all_members,
        'bandwidth':     bandwidth,
        'sprints':       Sprint.objects.filter(organization=org),
        'active_sprint': active_sprint,
        'org':           org,
        'subscription':  getattr(org, 'subscription', None),
    })


# ─────────────────────────────────────────
# MEMBER MANAGEMENT
# ─────────────────────────────────────────

@require_POST
@require_scrum_master_api
def add_member(request):
    from django.contrib.auth.models import User as AuthUser
    org       = get_org(request)
    data      = json.loads(request.body)
    username  = data.get('username', '').strip()
    stream_id = data.get('stream_id')
    role      = data.get('role', 'voter')

    if role not in ('admin', 'scrum_master', 'voter', 'viewer'):
        return JsonResponse({'error': 'Invalid role'}, status=400)

    # Only admins can assign admin role
    from .permissions import is_admin
    if role == 'admin' and not is_admin(request.user, org):
        return JsonResponse({'error': 'Only admins can assign admin role'}, status=403)

    try:
        user = AuthUser.objects.get(username=username)
    except AuthUser.DoesNotExist:
        return JsonResponse({'error': f'User "{username}" not found'}, status=400)

    stream = None
    if stream_id:
        stream = get_object_or_404(Stream, id=stream_id, organization=org)

    # Add or update OrganizationMember
    org_member, _ = OrganizationMember.objects.update_or_create(
        organization=org, user=user,
        defaults={'role': role}
    )

    # Add or update SprintMember
    member, created = SprintMember.objects.get_or_create(
        organization=org, user=user,
        defaults={'stream': stream, 'is_active': True}
    )
    if not created:
        member.stream    = stream
        member.is_active = True
        member.save()

    return JsonResponse({
        'ok':   True,
        'id':   member.id,
        'name': member.display_name(),
        'role': role,
    })


@require_POST
@require_scrum_master_api
def remove_member(request, member_id):
    org    = get_org(request)
    member = get_object_or_404(SprintMember, id=member_id, organization=org)

    # Prevent removing the last admin
    from .permissions import is_admin
    target_org_member = OrganizationMember.objects.filter(
        organization=org, user=member.user
    ).first()
    if target_org_member and target_org_member.role == 'admin':
        admin_count = OrganizationMember.objects.filter(
            organization=org, role='admin'
        ).count()
        if admin_count <= 1:
            return JsonResponse(
                {'error': 'Cannot remove the last admin of the organization'},
                status=400
            )

    member.is_active = False
    member.save()
    return JsonResponse({'ok': True})


@require_POST
@require_scrum_master_api
def change_member_role(request, member_id):
    """Change a member's role. Only admins can assign/remove admin role."""
    from .permissions import is_admin
    org    = get_org(request)
    member = get_object_or_404(SprintMember, id=member_id, organization=org)
    data   = json.loads(request.body)
    role   = data.get('role')

    if role not in ('admin', 'scrum_master', 'voter', 'viewer'):
        return JsonResponse({'error': 'Invalid role'}, status=400)

    # Only admins can assign or remove admin role
    target_org_member = get_object_or_404(
        OrganizationMember, organization=org, user=member.user
    )
    if (role == 'admin' or target_org_member.role == 'admin') and not is_admin(request.user, org):
        return JsonResponse({'error': 'Only admins can change admin role'}, status=403)

    # Prevent removing last admin
    if target_org_member.role == 'admin' and role != 'admin':
        admin_count = OrganizationMember.objects.filter(
            organization=org, role='admin'
        ).count()
        if admin_count <= 1:
            return JsonResponse(
                {'error': 'Cannot remove the last admin of the organization'},
                status=400
            )

    target_org_member.role = role
    target_org_member.save()
    return JsonResponse({'ok': True, 'role': role})


# ─────────────────────────────────────────
# STREAM MANAGEMENT
# ─────────────────────────────────────────

@require_POST
@require_admin_api
def add_stream(request):
    org  = get_org(request)
    data = json.loads(request.body)
    name = data.get('name', '').strip()
    if not name:
        return JsonResponse({'error': 'Stream name required'}, status=400)
    stream, _ = Stream.objects.get_or_create(organization=org, name=name)
    return JsonResponse({'ok': True, 'id': stream.id, 'name': stream.name})


@require_POST
@require_admin_api
def delete_stream(request, stream_id):
    org    = get_org(request)
    stream = get_object_or_404(Stream, id=stream_id, organization=org)
    stream.delete()
    return JsonResponse({'ok': True})


# ─────────────────────────────────────────
# SPRINT MANAGEMENT
# ─────────────────────────────────────────

@require_POST
@require_scrum_master_api
def add_sprint(request):
    org  = get_org(request)
    data = json.loads(request.body)
    name = data.get('name', '').strip()
    if not name:
        return JsonResponse({'error': 'Sprint name required'}, status=400)
    sprint = Sprint.objects.create(
        organization=org,
        name=name,
        goal=data.get('goal', ''),
        start_date=data.get('start_date') or None,
        end_date=data.get('end_date') or None,
        is_active=data.get('is_active', False),
    )
    if sprint.is_active:
        Sprint.objects.filter(organization=org).exclude(id=sprint.id).update(is_active=False)
    return JsonResponse({'ok': True, 'id': sprint.id, 'name': sprint.name})


@require_POST
@require_scrum_master_api
def edit_sprint(request, sprint_id):
    org    = get_org(request)
    sprint = get_object_or_404(Sprint, id=sprint_id, organization=org)
    data   = json.loads(request.body)
    for field in ['name', 'goal']:
        if field in data:
            setattr(sprint, field, data[field])
    for field in ['start_date', 'end_date']:
        if field in data:
            setattr(sprint, field, data[field] or None)
    if 'is_active' in data:
        sprint.is_active = data['is_active']
        if sprint.is_active:
            Sprint.objects.filter(organization=org).exclude(id=sprint.id).update(is_active=False)
    sprint.save()
    return JsonResponse({'ok': True})


@require_POST
@require_scrum_master_api
def delete_sprint(request, sprint_id):
    org    = get_org(request)
    sprint = get_object_or_404(Sprint, id=sprint_id, organization=org)
    sprint.delete()
    return JsonResponse({'ok': True})


# ─────────────────────────────────────────
# STORY MANAGEMENT
# ─────────────────────────────────────────

@require_POST
@require_scrum_master_api
def add_story(request):
    org   = get_org(request)
    data  = json.loads(request.body)
    title = data.get('title', '').strip()
    if not title:
        return JsonResponse({'error': 'Title required'}, status=400)
    owner  = get_object_or_404(SprintMember, id=data['owner_id'], organization=org) if data.get('owner_id') else None
    sprint = get_object_or_404(Sprint, id=data['sprint_id'], organization=org) if data.get('sprint_id') else None
    story  = UserStory.objects.create(
        organization=org,
        title=title,
        description=data.get('description', ''),
        owner=owner,
        sprint=sprint,
        involved_streams=data.get('involved_streams', []),
        order=UserStory.objects.filter(organization=org).count()
    )
    return JsonResponse({'ok': True, 'id': story.id})


@require_POST
@require_scrum_master_api
def edit_story(request, us_id):
    org   = get_org(request)
    story = get_object_or_404(UserStory, id=us_id, organization=org)
    data  = json.loads(request.body)
    if 'title'            in data: story.title            = data['title']
    if 'description'      in data: story.description      = data['description']
    if 'involved_streams' in data: story.involved_streams = data['involved_streams']
    if 'final_sp'         in data: story.final_sp         = data['final_sp']
    if 'owner_id'         in data:
        story.owner = get_object_or_404(
            SprintMember, id=data['owner_id'], organization=org
        ) if data['owner_id'] else None
    if 'sprint_id'        in data:
        story.sprint = get_object_or_404(
            Sprint, id=data['sprint_id'], organization=org
        ) if data['sprint_id'] else None
    story.save()
    return JsonResponse({'ok': True})


@require_POST
@require_scrum_master_api
def delete_story(request, us_id):
    org   = get_org(request)
    story = get_object_or_404(UserStory, id=us_id, organization=org)
    story.delete()
    return JsonResponse({'ok': True})


@require_POST
@require_scrum_master_api
def trigger_voting(request, us_id):
    org   = get_org(request)
    story = get_object_or_404(UserStory, id=us_id, organization=org)
    story.voting_status = 'voting'
    story.votes.all().delete()
    story.vote_average  = None
    story.save()
    return JsonResponse({'ok': True})


@require_POST
@require_scrum_master_api
def close_voting(request, us_id):
    org   = get_org(request)
    story = get_object_or_404(UserStory, id=us_id, organization=org)
    story.voting_status = 'closed'
    story.vote_average  = story.compute_average()
    story.save()
    return JsonResponse({'ok': True, 'average': story.vote_average})


@require_POST
@require_scrum_master_api
def assign_sp(request, us_id):
    org   = get_org(request)
    story = get_object_or_404(UserStory, id=us_id, organization=org)
    data  = json.loads(request.body)
    if 'final_sp' in data:
        story.final_sp = data['final_sp']
        story.save()
    if 'stream_assignments' in data:
        story.stream_assignments.all().delete()
        for sa in data['stream_assignments']:
            member = get_object_or_404(SprintMember, id=sa['member_id'], organization=org)
            stream = get_object_or_404(Stream, id=sa['stream_id'], organization=org)
            StreamAssignment.objects.create(
                user_story=story, stream=stream, member=member, sp=sa['sp']
            )
    return JsonResponse({'ok': True})


@require_POST
@require_scrum_master_api
def edit_stream_assignment(request, us_id):
    org  = get_org(request)
    data = json.loads(request.body)
    sa   = get_object_or_404(StreamAssignment, id=data['assignment_id'])
    sa.sp = data['sp']
    sa.save()
    return JsonResponse({'ok': True})


def get_story_detail(request, us_id):
    story       = get_object_or_404(UserStory, id=us_id)
    assignments = []
    for sa in story.stream_assignments.select_related('member', 'stream').all():
        assignments.append({
            'id':          sa.id,
            'stream_id':   sa.stream.id,
            'stream':      sa.stream.name,
            'member_id':   sa.member_id,
            'member_name': sa.member.display_name(),
            'sp':          sa.sp,
        })
    return JsonResponse({
        'id':                story.id,
        'title':             story.title,
        'description':       story.description,
        'owner_id':          story.owner_id,
        'owner_name':        story.owner.display_name() if story.owner else None,
        'involved_streams':  story.involved_streams,
        'final_sp':          story.final_sp,
        'voting_status':     story.voting_status,
        'vote_average':      story.vote_average,
        'sprint_id':         story.sprint_id,
        'stream_assignments': assignments,
    })


# ─────────────────────────────────────────
# EXPORT / IMPORT
# ─────────────────────────────────────────

@require_scrum_master
def export_sprint(request, sprint_id):
    org    = get_org(request)
    sprint = get_object_or_404(Sprint, id=sprint_id, organization=org)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse

    stories = UserStory.objects.filter(sprint=sprint).prefetch_related(
        'stream_assignments__member',
        'stream_assignments__stream',
        'votes__member'
    ).select_related('owner')

    wb  = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = 'User Stories'

    header_fill = PatternFill('solid', fgColor='4F46E5')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    alt_fill    = PatternFill('solid', fgColor='F1F0FF')
    border      = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    headers    = ['#', 'User Story', 'Description', 'Owner', 'Owner Stream',
                  'Involved Streams', 'Vote Average', 'Final SP', 'Voting Status']
    col_widths = [5, 45, 35, 20, 14, 30, 14, 10, 16]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws1.merge_cells('A1:I1')
    ws1['A1']           = f'Sprint: {sprint.name}'
    ws1['A1'].font      = Font(bold=True, size=13, color='1E1B4B')
    ws1['A1'].fill      = PatternFill('solid', fgColor='EEF2FF')
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 28

    header_row = 3
    for col, h in enumerate(headers, 1):
        cell           = ws1.cell(row=header_row, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = border

    for i, story in enumerate(stories, 1):
        row  = header_row + i
        fill = alt_fill if i % 2 == 0 else None
        vals = [
            i, story.title, story.description or '',
            story.owner.display_name() if story.owner else '—',
            story.owner.stream.name if story.owner and story.owner.stream else '—',
            ', '.join(story.involved_streams) if story.involved_streams else '—',
            story.vote_average or '—',
            story.final_sp or '—',
            story.get_voting_status_display(),
        ]
        for col, val in enumerate(vals, 1):
            cell           = ws1.cell(row=row, column=col, value=val)
            cell.border    = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if fill:
                cell.fill = fill

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{sprint.name.replace(" ", "_")}_export.xlsx"'
    )
    wb.save(response)
    return response


@require_scrum_master
def import_stories(request, sprint_id):
    org    = get_org(request)
    sprint = get_object_or_404(Sprint, id=sprint_id, organization=org)

    if request.method == 'GET':
        return render(request, 'planner/import_stories.html', {
            'sprint':  sprint,
            'members': SprintMember.objects.filter(
                organization=org, is_active=True
            ).select_related('user', 'stream'),
            'streams': Stream.objects.filter(organization=org),
        })

    import openpyxl
    from io import BytesIO

    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        return JsonResponse({'error': 'No file uploaded'}, status=400)

    try:
        wb      = openpyxl.load_workbook(BytesIO(excel_file.read()), data_only=True)
        ws      = wb.active
        headers = [
            str(cell.value).strip().lower() if cell.value else ''
            for cell in ws[1]
        ]

        def find_col(candidates):
            for c in candidates:
                for i, h in enumerate(headers):
                    if c in h:
                        return i
            return None

        col_title       = find_col(['title', 'user story', 'story', 'name'])
        col_description = find_col(['description', 'desc', 'detail'])
        col_sp          = find_col(['sp', 'story point', 'points', 'estimate'])

        if col_title is None:
            return JsonResponse({'error': 'Could not find a title column.'}, status=400)

        created = skipped = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            title = row[col_title] if col_title < len(row) else None
            if not title or str(title).strip() == '':
                skipped += 1
                continue
            final_sp = None
            if col_sp is not None and col_sp < len(row):
                try:
                    final_sp = float(row[col_sp]) if row[col_sp] is not None else None
                except (ValueError, TypeError):
                    pass
            UserStory.objects.create(
                organization=org,
                title=str(title).strip(),
                description=(
                    str(row[col_description]).strip()
                    if col_description and col_description < len(row) and row[col_description]
                    else ''
                ),
                sprint=sprint,
                final_sp=final_sp,
                order=UserStory.objects.filter(organization=org).count()
            )
            created += 1

        return JsonResponse({
            'ok':          True,
            'created':     created,
            'skipped':     skipped,
            'sprint_name': sprint.name,
        })
    except Exception as e:
        return JsonResponse({'error': f'Failed to read file: {str(e)}'}, status=400)
